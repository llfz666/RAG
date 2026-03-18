"""Integration tests for TableSplitter with DocumentChunker and XlsxLoader.

This module tests the end-to-end flow:
1. XlsxLoader loads Excel file and converts to Markdown tables
2. DocumentChunker auto-selects TableSplitter for .xlsx files
3. TableSplitter splits by row, preserving header in each chunk
4. BM25 retrieval can find complete row data
"""

import pytest
from pathlib import Path
from src.core.settings import load_settings
from src.core.types import Document
from src.ingestion.chunking.document_chunker import DocumentChunker
from src.libs.loader.xlsx_loader import XlsxLoader


@pytest.fixture
def settings():
    """Load application settings."""
    return load_settings()


@pytest.fixture
def sample_xlsx_file(project_root) -> Path:
    """Return path to a sample Excel file for testing."""
    # Try to find any xlsx file in fixtures
    fixtures_dir = project_root / "tests" / "fixtures"
    
    # Look for xlsx files
    for xlsx_file in fixtures_dir.rglob("*.xlsx"):
        return xlsx_file
    for xlsm_file in fixtures_dir.rglob("*.xlsm"):
        return xlsm_file
    
    # If no file found, create a temporary test file
    test_file = fixtures_dir / "test_table_splitter.xlsx"
    if not test_file.exists():
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestData"
        
        # Add header row
        ws['A1'] = '投保人'
        ws['B1'] = '保单号'
        ws['C1'] = '产品类型'
        ws['D1'] = '状态'
        ws['E1'] = '生效日期'
        
        # Add data rows
        test_data = [
            ('投保人 897', 'POL2026XXX', '健康保险', '有效', '2024-01-24'),
            ('投保人 587', 'POL2026YYY', '人寿保险', '有效', '2024-02-15'),
            ('投保人 398', 'POL2026ZZZ', '意外保险', '有效', '2024-03-10'),
        ]
        
        for i, row in enumerate(test_data, start=2):
            ws[f'A{i}'] = row[0]
            ws[f'B{i}'] = row[1]
            ws[f'C{i}'] = row[2]
            ws[f'D{i}'] = row[3]
            ws[f'E{i}'] = row[4]
        
        wb.save(test_file)
    
    return test_file


class TestDocumentChunkerAutoSelect:
    """Test DocumentChunker auto-selects TableSplitter for Excel files."""
    
    def test_auto_select_table_splitter_for_xlsx(self, settings, sample_xlsx_file):
        """Test that .xlsx files use TableSplitter."""
        chunker = DocumentChunker(settings)
        
        # Create a document with xlsx source_path
        doc = Document(
            id="test_xlsx_doc",
            text="| 投保人 | 保单号 |\n|----------|----------|\n| 投保人 897 | POL123 |",
            metadata={"source_path": str(sample_xlsx_file)}
        )
        
        # Get the splitter that would be used
        splitter = chunker._get_splitter_for_document(doc)
        
        # Should be TableSplitter
        from src.libs.splitter.table_splitter import TableSplitter
        assert isinstance(splitter, TableSplitter), \
            f"Expected TableSplitter for .xlsx file, got {type(splitter).__name__}"
    
    def test_auto_select_table_splitter_for_xlsm(self, settings):
        """Test that .xlsm files use TableSplitter."""
        chunker = DocumentChunker(settings)
        
        doc = Document(
            id="test_xlsm_doc",
            text="| 投保人 | 保单号 |\n|----------|----------|\n| 投保人 897 | POL123 |",
            metadata={"source_path": "/path/to/file.xlsm"}
        )
        
        splitter = chunker._get_splitter_for_document(doc)
        
        from src.libs.splitter.table_splitter import TableSplitter
        assert isinstance(splitter, TableSplitter)
    
    def test_use_recursive_splitter_for_pdf(self, settings):
        """Test that .pdf files use RecursiveSplitter (default)."""
        chunker = DocumentChunker(settings)
        
        doc = Document(
            id="test_pdf_doc",
            text="Some PDF content without table structure.",
            metadata={"source_path": "/path/to/file.pdf"}
        )
        
        splitter = chunker._get_splitter_for_document(doc)
        
        from src.libs.splitter.recursive_splitter import RecursiveSplitter
        assert isinstance(splitter, RecursiveSplitter)
    
    def test_force_splitter_override(self, settings):
        """Test force_splitter parameter for testing."""
        chunker = DocumentChunker(settings, force_splitter="table")
        
        doc = Document(
            id="test_doc",
            text="| Header |\n|--------|\n| Data |",
            metadata={"source_path": "/path/to/file.txt"}  # Not Excel extension
        )
        
        # Should use forced table splitter
        splitter = chunker._get_splitter_for_document(doc)
        
        from src.libs.splitter.table_splitter import TableSplitter
        assert isinstance(splitter, TableSplitter)


class TestXlsxLoaderWithTableSplitter:
    """Test XlsxLoader output works correctly with TableSplitter."""
    
    def test_xlsx_loader_produces_markdown_table(self, sample_xlsx_file):
        """Test that XlsxLoader produces Markdown table format."""
        loader = XlsxLoader()
        doc = loader.load(str(sample_xlsx_file))  # Returns single Document
        
        assert doc is not None
        
        # Should contain Markdown table syntax
        assert "|" in doc.text
        # Should have table separator row pattern
        assert "|---" in doc.text or "|:-" in doc.text or "| -" in doc.text
    
    def test_end_to_end_xlsx_to_chunks(self, settings, sample_xlsx_file):
        """Test full flow: XlsxLoader -> DocumentChunker -> Chunks."""
        # Step 1: Load Excel file
        loader = XlsxLoader()
        doc = loader.load(str(sample_xlsx_file))  # Returns single Document
        
        assert doc is not None
        
        # Step 2: Chunk with auto-selected splitter
        chunker = DocumentChunker(settings)
        chunks = chunker.split_document(doc)
        
        assert len(chunks) >= 1
        
        # Verify chunks have proper metadata
        for chunk in chunks:
            assert "chunk_index" in chunk.metadata
            assert "source_ref" in chunk.metadata
            assert chunk.metadata["source_ref"] == doc.id


class TestTableRowIntegrity:
    """Test that table row data integrity is preserved after splitting."""
    
    def test_row_data_not_split_across_chunks(self, settings):
        """Test that a single row's data stays in one chunk."""
        chunker = DocumentChunker(settings, force_splitter="table")
        
        # Create document with Markdown table
        table_content = """| 投保人 | 保单号 | 产品类型 | 状态 | 生效日期 |
|----------|----------|----------|----------|----------|
| 投保人 897 | POL2026XXX | 健康保险 | 有效 | 2024-01-24 |
| 投保人 587 | POL2026YYY | 人寿保险 | 有效 | 2024-02-15 |"""
        
        doc = Document(
            id="test_table",
            text=table_content,
            metadata={"source_path": "test.xlsx"}
        )
        
        chunks = chunker.split_document(doc)
        
        # Should have 2 chunks (one per data row)
        assert len(chunks) == 2
        
        # First chunk should have complete first row
        assert "投保人 897" in chunks[0].text
        assert "POL2026XXX" in chunks[0].text
        assert "健康保险" in chunks[0].text
        # First chunk should NOT have second row data
        assert "投保人 587" not in chunks[0].text
        
        # Second chunk should have complete second row
        assert "投保人 587" in chunks[1].text
        assert "POL2026YYY" in chunks[1].text
        assert "人寿保险" in chunks[1].text
        # Second chunk should NOT have first row data
        assert "投保人 897" not in chunks[1].text
    
    def test_each_chunk_has_header_for_context(self, settings):
        """Test that each chunk has header for self-contained context."""
        chunker = DocumentChunker(settings, force_splitter="table")
        
        table_content = """| 投保人 | 保单号 | 产品类型 |
|----------|----------|----------|
| A | P001 | 保险 A |
| B | P002 | 保险 B |"""
        
        doc = Document(
            id="test_table",
            text=table_content,
            metadata={"source_path": "test.xlsx"}
        )
        
        chunks = chunker.split_document(doc)
        
        # Each chunk should have header columns
        for chunk in chunks:
            assert "| 投保人 |" in chunk.text
            assert "| 保单号 |" in chunk.text
            assert "| 产品类型 |" in chunk.text


class TestBM25RetrievalWithTableChunks:
    """Test BM25 retrieval works correctly with table chunks."""
    
    def test_query_finds_complete_row(self, settings):
        """Test that BM25 query finds complete row data."""
        from src.ingestion.storage.bm25_indexer import BM25Indexer
        from src.ingestion.embedding.sparse_encoder import SparseEncoder
        
        chunker = DocumentChunker(settings, force_splitter="table")
        
        table_content = """| 投保人 | 保单号 | 产品类型 | 状态 |
|----------|----------|----------|----------|
| 投保人 897 | POL2026XXX | 健康保险 | 有效 |
| 投保人 587 | POL2026YYY | 人寿保险 | 有效 |
| 投保人 398 | POL2026ZZZ | 意外保险 | 有效 |"""
        
        doc = Document(
            id="test_table",
            text=table_content,
            metadata={"source_path": "test.xlsx"}
        )
        
        chunks = chunker.split_document(doc)
        
        # Build BM25 index using SparseEncoder output
        # SparseEncoder.encode() expects List[Chunk] objects, not List[str]
        sparse_encoder = SparseEncoder()
        term_stats = sparse_encoder.encode(chunks)  # Pass Chunk objects directly
        
        indexer = BM25Indexer()
        indexer.build(term_stats, collection="test_table")
        
        # Query for specific policy holder
        results = indexer.query(["投保人", "897"], top_k=1)
        
        assert len(results) >= 1
        top_chunk_id = results[0]["chunk_id"]
        
        # Find the chunk text by id
        chunk_text_map = {c.id: c.text for c in chunks}
        top_result_text = chunk_text_map.get(top_chunk_id, "")
        
        # Should find the chunk with 投保人 897
        assert "投保人 897" in top_result_text
        # Should also have complete row data
        assert "POL2026XXX" in top_result_text
        assert "健康保险" in top_result_text
